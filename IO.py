"""
2021-05-22 이범석 21:50 엑셀파일 내 플레이기록과 진행사항 저장 함수 작성
2021-05-27 이범석 날짜별 플레이 타임 저장 함수 작성, 전체적인 코드 수정
2021-06-12 이범석 저장된 기록을 출력할 함수 작성
"""
import openpyxl
from datetime import datetime
from openpyxl import Workbook


MAX_ROW = 8  # 총 도전과제 수 +1
MAX_COLUMN = 11  # 총 음원 수 +1

# 음원 리스트와 도전과제 리스트
SList = ['낮의 달', '스타쉽 몽상', '추적자', 'swing_swing', 'Space Town', 'New_Future', 'Funky_Junky', 'Film_Clash!',
         'Feather_of_the_Angel', 'Expantion']
TimeList = [69, 142, 309, 193, 95, 181, 165, 86, 73, 188]


# 초기 엑셀파일 생성
def Init_file():
    # 유저 진행정보 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet'

    # 음원 리스트 입력, 누적 횟수 초기화
    for r in range(2, MAX_COLUMN+1):
        ws.cell(row=1, column=r, value=SList[r-2])
        ws.cell(row=2, column=r, value=0)
    # ROW값 입력, 날짜 초기화
    today = datetime.now()
    ws.cell(row=2, column=1, value='누적횟수')
    ws.cell(row=4, column=1, value='헬창')
    ws.cell(row=5, column=1, value='총 횟수')
    ws.cell(row=4, column=2, value=today)
    ws.cell(row=4, column=3, value=today)
    # 총 횟수 초기화
    ws.cell(row=5, column=2, value=0)
    # 금일 플레이타임 초기화
    ws.cell(row=6, column=1, value='금일 플레이타임')
    ws.cell(row=6, column=2, value=0)

    wb.save('UserData/user.xlsx')
    wb.close()

    # 금일 플레이 기록 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    ws.cell(row=1, column=1, value='음원')
    ws.cell(row=1, column=2, value='정확도')
    ws.cell(row=1, column=3, value='날짜')

    wb.save('UserData/accuracy.xlsx')
    wb.close()

    # 날짜별 운동시간 기록 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet2'

    ws.cell(row=1, column=1, value='날짜')
    ws.cell(row=2, column=1, value=datetime.now())
    ws.cell(row=1, column=2, value='플레이타임')
    ws.cell(row=2, column=2, value=0)
    ws.cell(row=1, column=3, value='음원(곡)')
    ws.cell(row=2, column=3, value=0)

    wb.save('UserData/count.xlsx')
    wb.close()


# 금일 플레이 기록 출력
def Print_today():
    wb = openpyxl.load_workbook('UserData/accuracy.xlsx', data_only=True)
    ws = wb.active
    ws.title = 'Sheet1'

    for r in range(2, ws.max_row+1):
        name = ws.cell(row=r, column=1).value
        accuracy = ws.cell(row=r, column=2).value
        print(str(name)+": "+str(accuracy))

    wb.close()


# 유저의 게임 진행정보를 변경
def Play_count(song):
    global date_diff, total
    for r in range(0, MAX_COLUMN - 1):
        sum = 0
        # 노래의 index값을 찾은 경우
        if song == SList[r]:
            wb = openpyxl.load_workbook('UserData/user.xlsx', data_only=True)
            ws = wb['Sheet']
            # 완료한 음원의 누적횟수 1회증가
            B = ws.cell(row=2, column=r + 2).value
            ws.cell(row=2, column=r + 2, value=B + 1)

            # 금일 플레이 타임 계산
            playtime = ws.cell(row=6, column=2).value
            playtime += TimeList[r]
            ws.cell(row=6, column=2, value=playtime)

            # 누적횟수를 모두 더하여 총횟수를 게산
            for r in range(2, MAX_COLUMN):
                sum += ws.cell(row=2, column=r).value

            ws.cell(row=5, column=2, value=sum)

            # 도전과제 '헬창'을 위한 날짜 계산
            last = ws.cell(row=4, column=3).value  # 최근 플레이 날짜
            today = datetime.now()  # 현재 날짜
            date_diff = today - last
            total = (today - ws.cell(row=4, column=2).value).days
            Cumul = ws.cell(row=5, column=2).value

            # 날짜 변경에 따른 금일 운동시간 초기화
            if date_diff.days != 0:
                ws.cell(row=6, column=2, value=0)

            # 운동을 하루라도 쉰 경우
            if date_diff.days != 0 and date_diff.days != 1:
                ws.cell(row=4, column=2, value=today)
                ws.cell(row=4, column=3, value=today)

            wb.save('UserData/user.xlsx')
            wb.close()

    if date_diff.days != 0 and date_diff.days != 1:
        return 0
    else:
        return total, Cumul


# 하루 지난 플레이 기록을 삭제
def Del_record():
    wb = openpyxl.load_workbook('UserData/accuracy.xlsx', data_only=True)
    ws = wb['Sheet1']
    now = datetime.now()
    for r in range(2, ws.max_row + 1):
        time = ws.cell(row=r, column=3).value
        compare = (now - time).days
        if compare != 0:
            ws.delete_rows(r)
        else:
            pass
    wb.save('UserData/accuracy.xlsx')
    wb.close()


# 플레이 정보를 기록
def Play_record(song, accuracy):
    wb = openpyxl.load_workbook('UserData/user.xlsx', data_only=True)
    ws = wb['Sheet']
    playtime = ws.cell(row=6, column=2).value
    wb.close()

    wb = openpyxl.load_workbook('UserData/accuracy.xlsx', data_only=True)
    ws = wb['Sheet1']
    today = datetime.now()
    ws.append([song, accuracy, today])
    today_play = ws.max_row-1

    wb.save('UserData/accuracy.xlsx')
    wb.close()

    wb = openpyxl.load_workbook('UserData/count.xlsx', data_only=True)
    ws = wb['Sheet2']
    time = ws.cell(row=ws.max_row, column=1).value
    date_diff = today - time
    if date_diff.days == 0:
        ws.cell(row=ws.max_row, column=1, value=today)
        ws.cell(row=ws.max_row, column=2, value=playtime)
        ws.cell(row=ws.max_row, column=3, value=today_play)
    if date_diff.days != 0:
        ws.cell(row=ws.max_row+1, column=1, value=today)
        ws.cell(row=ws.max_row+1, column=2, value=playtime)
        ws.cell(row=ws.max_row+1, column=3, value=today_play)

    wb.save('UserData/count.xlsx')
    wb.close()

# 테스트 함수
def main(song, accuracy):
    Play_count(song)
    Play_record(song, accuracy)
    """t, c = Play_count(song)
    print(c)
    if t == 100:
        print("헬창!")
    elif t == 50:
        print("에어로빅 강사")
    elif c == 1:
        print("시작이 반이다")
    else:
        pass"""
