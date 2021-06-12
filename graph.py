# plotly 라이브러리 불러오기
import plotly
import plotly.graph_objs as go
import openpyxl

wb = openpyxl.load_workbook('UserData/count.xlsx', data_only=True)
ws = wb['Sheet2']

date = [0]*(ws.max_row-1)
time = [0]*(ws.max_row-1)
for r in range(2, ws.max_row+1):
    date[r-2] = ws.cell(row=ws.max_row+2-r, column=1).value
    time[r-2] = ws.cell(row=ws.max_row+2-r, column=2).value
wb.close()

print(date)
print(time)
plotly.offline.plot({
    "data": [go.Scatter(x=date, y=time)],
    "layout": go.Layout(title="Time of exercise")
}, auto_open=True)
